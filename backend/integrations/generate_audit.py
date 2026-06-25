from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

# ── style constants ───────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
ALT_FILL     = PatternFill("solid", start_color="D6E4F0")
TOTAL_FILL   = PatternFill("solid", start_color="E2EFDA")
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(name="Arial", size=10)
TOTAL_FONT   = Font(name="Arial", bold=True, size=10)
TITLE_FONT   = Font(name="Arial", bold=True, size=13, color="1F4E79")
thin         = Side(style="thin", color="B0B0B0")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM_FMT      = '#,##0'

def _h(cell):
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def _b(cell, alt=False, center=False):
    cell.font      = BODY_FONT
    cell.fill      = ALT_FILL if alt else WHITE_FILL
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    cell.border    = BORDER

def _n(cell, alt=False):
    cell.font         = BODY_FONT
    cell.fill         = ALT_FILL if alt else WHITE_FILL
    cell.alignment    = Alignment(horizontal="right", vertical="center")
    cell.border       = BORDER
    cell.number_format = NUM_FMT

def _t(cell):
    cell.font         = TOTAL_FONT
    cell.fill         = TOTAL_FILL
    cell.alignment    = Alignment(horizontal="right", vertical="center")
    cell.border       = BORDER
    cell.number_format = NUM_FMT

def _title(ws, text, span, period_label):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, text)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c2 = ws.cell(2, 1, f"Period: {period_label}   |   Generated: {date.today().strftime('%d %B %Y')}")
    c2.font = Font(name="Arial", size=9, italic=True, color="666666")
    c2.alignment = Alignment(horizontal="left")
    ws.row_dimensions[3].height = 6  # spacer

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _fmt_date(d):
    if not d:
        return ""
    return str(d)

# ── sheet: Suppliers ──────────────────────────────────────────────────────────
def build_suppliers(wb, suppliers, period_label):
    ws = wb.create_sheet("Suppliers")
    _title(ws, "Supplier Purchases", 5, period_label)

    for col, h in enumerate(["Date Received", "Supplier Name", "Item Type / Cylinder Size", "Qty Received", "Amount Paid (NPR)"], 1):
        _h(ws.cell(4, col, h))

    grand_qty = grand_amt = 0
    for i, s in enumerate(suppliers):
        alt = i % 2 == 1
        r   = i + 5
        size = f"{s.get('item_type','')} / {s.get('cylinder_size','')}" if s.get('cylinder_size') else s.get('item_type','')
        _b(ws.cell(r, 1, _fmt_date(s.get("date_received"))),  alt, center=True)
        _b(ws.cell(r, 2, s.get("supplier_name", "")),         alt)
        _b(ws.cell(r, 3, size),                               alt)
        _n(ws.cell(r, 4, s.get("quantity_received", 0)),      alt)
        _n(ws.cell(r, 5, s.get("amount_paid", 0) or 0),       alt)
        grand_qty += s.get("quantity_received", 0) or 0
        grand_amt += s.get("amount_paid", 0) or 0

    tr = len(suppliers) + 5
    ws.cell(tr, 1).border = BORDER; ws.cell(tr, 1).fill = TOTAL_FILL
    ws.cell(tr, 2).border = BORDER; ws.cell(tr, 2).fill = TOTAL_FILL
    ws.cell(tr, 3, "TOTAL").font = TOTAL_FONT; ws.cell(tr, 3).fill = TOTAL_FILL; ws.cell(tr, 3).border = BORDER
    ws.cell(tr, 3).alignment = Alignment(horizontal="right")
    _t(ws.cell(tr, 4, f"=SUM(D5:D{tr-1})"))
    _t(ws.cell(tr, 5, f"=SUM(E5:E{tr-1})"))

    _widths(ws, [16, 28, 28, 16, 20])
    ws.row_dimensions[4].height = 22

# ── sheet: Customers ──────────────────────────────────────────────────────────
def build_customers(wb, customers, period_label):
    ws = wb.create_sheet("Customers")
    _title(ws, "Customer Deposits & Returns", 7, period_label)

    for col, h in enumerate(["Deposit Date", "Customer Name", "Phone", "Item Type / Size", "Qty", "Deposit Amount (NPR)", "Returned Date"], 1):
        _h(ws.cell(4, col, h))

    for i, c in enumerate(customers):
        alt  = i % 2 == 1
        r    = i + 5
        size = f"{c.get('item_type','')} / {c.get('cylinder_size','')}" if c.get('cylinder_size') else c.get('item_type','')
        _b(ws.cell(r, 1, _fmt_date(c.get("deposit_date"))),   alt, center=True)
        _b(ws.cell(r, 2, c.get("full_name", "")),             alt)
        _b(ws.cell(r, 3, c.get("phone", "")),                 alt)
        _b(ws.cell(r, 4, size),                               alt)
        _n(ws.cell(r, 5, c.get("quantity", 0)),               alt)
        _n(ws.cell(r, 6, c.get("deposit_amount", 0) or 0),    alt)
        _b(ws.cell(r, 7, _fmt_date(c.get("returned_date")) or "Not returned"), alt, center=True)

    tr = len(customers) + 5
    for col in [1, 2, 3, 4, 7]:
        ws.cell(tr, col).fill = TOTAL_FILL; ws.cell(tr, col).border = BORDER
    ws.cell(tr, 4, "TOTAL").font = TOTAL_FONT; ws.cell(tr, 4).fill = TOTAL_FILL
    ws.cell(tr, 4).alignment = Alignment(horizontal="right"); ws.cell(tr, 4).border = BORDER
    _t(ws.cell(tr, 5, f"=SUM(E5:E{tr-1})"))
    _t(ws.cell(tr, 6, f"=SUM(F5:F{tr-1})"))

    _widths(ws, [16, 25, 16, 28, 10, 22, 16])
    ws.row_dimensions[4].height = 22

# ── sheet: Cylinder Sales (refill) ────────────────────────────────────────────
def build_cylinder_sales(wb, cylinder_sales, period_label):
    ws = wb.create_sheet("Cylinder Sales")
    _title(ws, "Cylinder Sales (Refill)", 5, period_label)

    for col, h in enumerate(["Sale Date", "Customer Name", "Phone", "Item Type", "Cylinder Size", "Qty Sold", "Amount (NPR)"], 1):
        _h(ws.cell(4, col, h))

    for i, s in enumerate(cylinder_sales):
        alt = i % 2 == 1
        r   = i + 5
        _b(ws.cell(r, 1, _fmt_date(s.get("sale_date"))),       alt, center=True)
        _b(ws.cell(r, 2, s.get("customer_name", "")),          alt)
        _b(ws.cell(r, 3, s.get("customer_phone", "")),         alt)
        _b(ws.cell(r, 4, s.get("item_type", "")),              alt)
        _b(ws.cell(r, 5, s.get("cylinder_size", "") or "—"),   alt)
        _n(ws.cell(r, 6, s.get("quantity", 0)),                alt)
        _n(ws.cell(r, 7, s.get("amount", 0) or 0),            alt)

    tr = len(cylinder_sales) + 5
    for col in [1, 2, 3, 4, 5]:
        ws.cell(tr, col).fill = TOTAL_FILL; ws.cell(tr, col).border = BORDER
    ws.cell(tr, 5, "TOTAL").font = TOTAL_FONT; ws.cell(tr, 5).fill = TOTAL_FILL
    ws.cell(tr, 5).alignment = Alignment(horizontal="right"); ws.cell(tr, 5).border = BORDER
    _t(ws.cell(tr, 6, f"=SUM(F5:F{tr-1})"))
    _t(ws.cell(tr, 7, f"=SUM(G5:G{tr-1})"))

    # column widths — add two more:
    _widths(ws, [16, 25, 16, 22, 20, 14, 20])

# ── sheet: Accessory Sales ────────────────────────────────────────────────────
def build_accessory_sales(wb, accessory_sales, period_label):
    ws = wb.create_sheet("Accessory Sales")
    _title(ws, "Accessory Sales", 4, period_label)

    for col, h in enumerate(["Sale Date", "Item Type (Accessory)", "Qty Sold", "Amount (NPR)"], 1):
        _h(ws.cell(4, col, h))

    for i, s in enumerate(accessory_sales):
        alt = i % 2 == 1
        r   = i + 5
        _b(ws.cell(r, 1, _fmt_date(s.get("sale_date"))),  alt, center=True)
        _b(ws.cell(r, 2, s.get("item_type", "")),         alt)
        _n(ws.cell(r, 3, s.get("quantity", 0)),           alt)
        _n(ws.cell(r, 4, s.get("amount", 0) or 0),        alt)

    tr = len(accessory_sales) + 5
    for col in [1, 2]:
        ws.cell(tr, col).fill = TOTAL_FILL; ws.cell(tr, col).border = BORDER
    ws.cell(tr, 2, "TOTAL").font = TOTAL_FONT; ws.cell(tr, 2).fill = TOTAL_FILL
    ws.cell(tr, 2).alignment = Alignment(horizontal="right"); ws.cell(tr, 2).border = BORDER
    _t(ws.cell(tr, 3, f"=SUM(C5:C{tr-1})"))
    _t(ws.cell(tr, 4, f"=SUM(D5:D{tr-1})"))

    _widths(ws, [16, 28, 14, 20])
    ws.row_dimensions[4].height = 22

# ── main ──────────────────────────────────────────────────────────────────────
def generate_audit_excel(suppliers, customers, cylinder_sales, accessory_sales, period_label="All"):
    wb = Workbook()
    wb.remove(wb.active)
    build_suppliers(wb, suppliers, period_label)
    build_customers(wb, customers, period_label)
    build_cylinder_sales(wb, cylinder_sales, period_label)
    build_accessory_sales(wb, accessory_sales, period_label)
    return wb


# ── smoke test ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    suppliers = [
        {"date_received": "2026-06-23", "supplier_name": "Nepal Gas Ltd", "item_type": "Nitrogen", "cylinder_size": "Large",  "quantity_received": 15, "amount_paid": 20000},
        {"date_received": "2026-06-23", "supplier_name": "Nepal Gas Ltd", "item_type": "Oxygen",   "cylinder_size": "Small",  "quantity_received": 10, "amount_paid": 12000},
    ]
    customers = [
        {"deposit_date": "2026-06-23", "full_name": "Ram Bahadur", "phone": "9800000001", "item_type": "Nitrogen", "cylinder_size": "Large",  "quantity": 5,  "deposit_amount": 12000, "returned_date": None},
        {"deposit_date": "2026-06-22", "full_name": "Sita Devi",   "phone": "9800000002", "item_type": "Oxygen",   "cylinder_size": "Small",  "quantity": 3,  "deposit_amount": 7500,  "returned_date": "2026-06-23"},
    ]
    cylinder_sales = [
        {"sale_date": "2026-06-23", "item_type": "Nitrogen", "cylinder_size": "Large",  "quantity": 2, "amount": 500},
        {"sale_date": "2026-06-23", "item_type": "Oxygen",   "cylinder_size": "Small",  "quantity": 1, "amount": 300},
    ]
    accessory_sales = [
        {"sale_date": "2026-06-23", "item_type": "Regulator", "quantity": 10, "amount": 10000},
        {"sale_date": "2026-06-23", "item_type": "Mask",      "quantity": 5,  "amount": 2500},
    ]
    wb = generate_audit_excel(suppliers, customers, cylinder_sales, accessory_sales, "Today (23 June 2026)")
    wb.save("/home/claude/audit_report2.xlsx")